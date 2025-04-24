import streamlit as st
import pandas as pd
import pyodbc
from io import BytesIO
from datetime import datetime
import re
import zipfile
from openpyxl import load_workbook
from xlrd.compdoc import CompDocError


def get_db_connection():
    # Recuperar as credenciais configuradas via secrets
    server   = st.secrets["mssql"]["server"]
    database = st.secrets["mssql"]["database"]
    username = st.secrets["mssql"]["username"]
    password = st.secrets["mssql"]["password"]
    if not all([server, database, username, password]):
        st.error("Verifique se todas as variáveis estão definidas corretamente no secrets.")
        st.stop()
    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"
    try:
        conn = pyodbc.connect(conn_str)
        return conn
    except Exception as e:
        st.error("Erro na conexão com o banco de dados: " + str(e))
        st.stop()

def run_system_query(start_date, end_date, id_empresa, client_ids, id_forma_list):
    conn = get_db_connection()
    # Formatação das datas
    start_str = start_date.strftime('%Y-%m-%d 00:00:00')
    end_str = end_date.strftime('%Y-%m-%d 23:59:59')
    client_ids_str = ",".join(str(x) for x in client_ids)
    id_forma_str = ",".join(str(x) for x in id_forma_list)
    
    query = f"""
    Select
        CA.ID_Venda,
        FP.Descricao [Forma de Pagamento],
        U.Nome,
        CA.ID_Caixa,
        FP.ID_Forma,
        CA.Documento_Cartao NSU,
        Replace(Convert(Varchar,Convert(Decimal(18,2),CA.Valor)),'.',',') [Valor Bruto],
        case when format(VS.Data_Faturamento, 'dd/MM/yyyy') is null 
             then format(CA.datacadastro, 'dd/MM/yyyy') 
             else format(VS.Data_Faturamento, 'dd/MM/yyyy') end as Data_Faturamento1,
        case when VS.Data_Faturamento is null 
             then CA.datacadastro 
             else VS.Data_Faturamento end as Data_Faturamento,
        c.ID_Cliente, 
        c.RazaoCliente
    From ContasAReceber CA
    left Join FormasPagamento FP On FP.ID_Forma = CA.ID_Forma
    inner Join Fechamento_Caixas FC On
            FC.ID_Empresa = CA.ID_Empresa And
            FC.ID_Caixa = CA.ID_Caixa and
            FC.ID_Origem_Caixa = CA.id_origem_caixa
    inner Join Usuarios U On U.ID_Usuario = FC.ID_Usuario
    left Join Vendas_Sorveteria VS On
            VS.ID_Empresa = CA.ID_Empresa And
            VS.ID_Venda = CA.ID_Venda
    Inner Join Empresas E On CA.ID_Empresa = E.ID_Empresa
    inner join Clientes C on CA.ID_Cliente = C.ID_Cliente
    Where CA.ID_Forma In (1,31,5,6,17,18,37)
      And E.TipoEmpresa = 'Sorveteria'
      And IsNull(CA.Emissao, VS.Data_Faturamento) >= '{start_str}'
      And IsNull(CA.Emissao, VS.Data_Faturamento) <= '{end_str}'
      and E.ID_Empresa in ({id_empresa})
      and ca.ID_Origem_Caixa = 1
      and c.ID_Cliente in ({client_ids_str})
      and FP.ID_Forma in ({id_forma_str})
    Order By FP.ID_Forma, CA.Valor, E.NomeFantasia, IsNull(VS.Data_Faturamento, CA.Emissao)
    """
    df = pd.read_sql(query, conn)
    # Converter a coluna de data e ajustar o valor bruto
    df['Data_Faturamento'] = pd.to_datetime(df['Data_Faturamento'], errors='coerce').dt.date
    df['Valor Bruto'] = df['Valor Bruto'].str.replace(',', '.', regex=False)
    df['Valor Bruto'] = pd.to_numeric(df['Valor Bruto'], errors='coerce').fillna(0)
    return df

def consolidate_data(delivery_df, system_df, date_col_delivery, value_col_delivery,
                     date_col_system, value_col_system, order_delivery_cols, order_system_cols):
    """
    Consolida os dados da planilha do delivery com os dados do sistema com base na correspondência
    por data e valor.
    order_delivery_cols e order_system_cols são dicionários que mapeiam o nome das colunas no output.
    """
    # Converter e normalizar as datas
    delivery_df[date_col_delivery] = pd.to_datetime(delivery_df[date_col_delivery], dayfirst=True, errors='coerce').dt.date
    system_df[date_col_system] = pd.to_datetime(system_df[date_col_system], errors='coerce').dt.date

    delivery_df.sort_values(by=[date_col_delivery, value_col_delivery], inplace=True)
    system_df.sort_values(by=[date_col_system, value_col_system], inplace=True)
    delivery_df.reset_index(drop=True, inplace=True)
    system_df.reset_index(drop=True, inplace=True)

    system_df['Matched'] = False
    results = []
    # Realiza a correspondência dos registros
    for i, d_row in delivery_df.iterrows():
        condition = (system_df[date_col_system] == d_row[date_col_delivery]) & \
                    (system_df[value_col_system] == d_row[value_col_delivery]) & \
                    (~system_df['Matched'])
        match = system_df[condition]
        if not match.empty:
            matched_row = match.iloc[0]
            results.append((d_row, matched_row, 'Correspondente'))
            system_df.loc[matched_row.name, 'Matched'] = True
        else:
            results.append((d_row, pd.Series(), 'Diferença'))
    for j, s_row in system_df.iterrows():
        if not s_row['Matched']:
            results.append((pd.Series(), s_row, 'Diferença'))
    
    consolidated_list = []
    for d_row, s_row, status in results:
        row_dict = {}
        # Preenche as colunas vindas da planilha do delivery
        for new_col, orig_col in order_delivery_cols.items():
            if orig_col in d_row and pd.notnull(d_row.get(orig_col)):
                if new_col.lower().find('data') != -1 and isinstance(d_row[orig_col], (pd.Timestamp, datetime)):
                    row_dict[new_col] = d_row[orig_col].strftime('%d/%m/%Y')
                else:
                    row_dict[new_col] = d_row[orig_col]
            else:
                row_dict[new_col] = ''
        # Preenche as colunas vindas do sistema
        for new_col, orig_col in order_system_cols.items():
            if orig_col in s_row and pd.notnull(s_row.get(orig_col)):
                if new_col.lower().find('data') != -1 and isinstance(s_row[orig_col], (pd.Timestamp, datetime)):
                    row_dict[new_col] = s_row[orig_col].strftime('%d/%m/%Y')
                else:
                    row_dict[new_col] = s_row[orig_col]
            else:
                row_dict[new_col] = ''
        row_dict['Discrepância Inicial'] = status
        consolidated_list.append(row_dict)
    
    final_df = pd.DataFrame(consolidated_list)
    return final_df

# Funções de processamento para cada delivery:

def process_ifood(uploaded_file):
    """
    Processa a planilha do IFood.
    Colunas esperadas: 'N° PEDIDO', 'DATA', 'VALOR DOS ITENS'
    """
    df = pd.read_excel(uploaded_file)
    required = ['N° PEDIDO', 'DATA', 'VALOR DOS ITENS']
    df = df[required]
    df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce').dt.date
    df['VALOR DOS ITENS'] = pd.to_numeric(df['VALOR DOS ITENS'], errors='coerce').fillna(0)
    df.rename(columns={'N° PEDIDO': 'N° PEDIDO IFOOD', 'DATA': 'DATA IFOOD', 'VALOR DOS ITENS': 'VALOR IFOOD'}, inplace=True)
    return df

def process_mais_delivery(uploaded_file):
    """
    Processa a planilha do Mais Delivery.
    Colunas esperadas: 'Data Pedido', 'Número', 'Valor (R$)'
    """
    df = pd.read_excel(uploaded_file)
    required = ['Data Pedido', 'Número', 'Valor (R$)']
    df = df[required]
    df['Data Pedido'] = pd.to_datetime(df['Data Pedido'], dayfirst=True, errors='coerce').dt.date
    # Usar raw strings para os padrões de regex
    df['Valor (R$)'] = df['Valor (R$)'].replace({r'R\$': '', ',': '.', r'\s+': ''}, regex=True)
    df['Valor (R$)'] = pd.to_numeric(df['Valor (R$)'], errors='coerce').fillna(0)
    df.rename(columns={'Data Pedido': 'DATA MAIS DELIVERY', 'Número': 'N° PEDIDO MAIS DELIVERY', 'Valor (R$)': 'VALOR MAIS DELIVERY'}, inplace=True)
    return df

def process_aiquefome(uploaded_file):
    """
    Processa a planilha do AI QUE FOME.
    Colunas esperadas: 'Nro. Pedido', 'Data', 'Total (R$)', 'Desconto (R$)'
    """
    try:
        df = pd.read_excel(uploaded_file)
    except CompDocError:
        st.error(
            "Não foi possível ler o arquivo AI QUE FOME (.xls). "
            "Ele parece corrompido ou num formato antigo. "
            "Abra no Excel e salve como .xlsx, depois tente novamente."
        )
        st.stop()
    # --- continua o tratamento normal ---
    required = ['Nro. Pedido', 'Data', 'Total (R$)', 'Desconto (R$)']
    df = df[required]
    df['Data'] = pd.to_datetime(df['Data'], dayfirst=True, errors='coerce').dt.date
    for col in ['Total (R$)', 'Desconto (R$)']:
        df[col] = df[col].replace({r'R\$': '', ',': '.', r'\s+': ''}, regex=True)
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Valor AI QUE FOME'] = df['Total (R$)'] + df['Desconto (R$)']
    df.rename(columns={'Nro. Pedido': 'N° PEDIDO AI QUE FOME', 'Data': 'DATA AI QUE FOME'}, inplace=True)
    return df


def process_goomer(uploaded_file):
    """
    Processa a planilha do GOOMER.
    Colunas esperadas: ['ID do pedido','Data', 'Cupom','Cupom (R$)', 'Total (R$)', 'Tipo', 'Forma de pagamento']
    """
    df = pd.read_excel(uploaded_file)
    required = ['ID do pedido','Data', 'Cupom','Cupom (R$)', 'Total (R$)', 'Tipo', 'Forma de pagamento']
    df = df[required]
    df['Data'] = pd.to_datetime(df['Data'], dayfirst=True, errors='coerce').dt.date
    df['Total (R$)'] = df['Total (R$)'].replace({r'R\$': '', ',': '.', r'\s+': ''}, regex=True)
    df['Total (R$)'] = pd.to_numeric(df['Total (R$)'], errors='coerce').fillna(0)
    df.rename(columns={
        'ID do pedido': 'N° PEDIDO GOOMER',
        'Data': 'DATA GOOMER',
        'Cupom': 'CUPOM',
        'Cupom (R$)': 'CUPOM (R$)',
        'Total (R$)': 'VALOR GOOMER',
        'Tipo': 'TIPO',
        'Forma de pagamento': 'FORMA DE PAGAMENTO'
    }, inplace=True)
    return df

def read_excel_nostyles(uploaded_file):
    """
    Lê o arquivo Excel ignorando os estilos problemáticos.
    Essa função lê o arquivo como um zip, substitui o conteúdo de 'xl/styles.xml'
    por um nó mínimo de fills e reconstroi o arquivo, garantindo que todas as linhas
    tenham o mesmo número de colunas, preenchendo com None as células ausentes.
    """
    # Garantir que temos um objeto BytesIO
    if not hasattr(uploaded_file, 'read'):
         uploaded_file = BytesIO(uploaded_file)
    else:
         uploaded_file.seek(0)
    content = uploaded_file.read()
    uploaded_file = BytesIO(content)
    try:
        with zipfile.ZipFile(uploaded_file, "r") as zin:
            filelist = zin.namelist()
            if "xl/styles.xml" in filelist:
                styles = zin.read("xl/styles.xml").decode("utf-8")
                # Substituir o nó <fills> por um nó mínimo válido
                styles = re.sub(r'<fills[^>]*>.*?</fills>', '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>', styles, flags=re.DOTALL)
                mem_zip = BytesIO()
                with zipfile.ZipFile(mem_zip, mode="w") as zout:
                    for item in zin.infolist():
                        if item.filename == "xl/styles.xml":
                            zout.writestr(item, styles)
                        else:
                            zout.writestr(item, zin.read(item.filename))
                mem_zip.seek(0)
                wb = load_workbook(mem_zip, read_only=True, data_only=True)
            else:
                uploaded_file.seek(0)
                wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as e:
        st.error("Erro ao ler planilha sem estilos: " + str(e))
        raise e
    ws = wb.active
    data = ws.values
    cols = list(next(data))
    expected_cols = len(cols)
    data_rows = []
    for row in data:
        row = list(row)
        if len(row) < expected_cols:
            row.extend([None]*(expected_cols - len(row)))
        data_rows.append(row)
    return pd.DataFrame(data_rows, columns=cols)

def process_edvania(uploaded_file):
    """
    Processa a planilha do EDVANIA SOBRINHO DA SILVA (Mais Delivery Guaraí).
    Utiliza a coluna "Total com entrega (R$)" em vez de "Valor (R$)".
    """
    df = read_excel_nostyles(uploaded_file)
    required = ['Data Pedido', 'Número', 'Total com entrega (R$)']
    df = df[required]
    df['Data Pedido'] = pd.to_datetime(df['Data Pedido'], dayfirst=True, errors='coerce').dt.date
    df['Total com entrega (R$)'] = df['Total com entrega (R$)'].replace({r'R\$': '', ',': '.', r'\s+': ''}, regex=True)
    df['Total com entrega (R$)'] = pd.to_numeric(df['Total com entrega (R$)'], errors='coerce').fillna(0)
    df.rename(columns={'Data Pedido': 'DATA MAIS DELIVERY', 'Número': 'N° PEDIDO MAIS DELIVERY', 
                         'Total com entrega (R$)': 'VALOR MAIS DELIVERY'}, inplace=True)
    return df
